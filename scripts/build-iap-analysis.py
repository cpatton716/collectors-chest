"""Build the Native App IAP Analysis workbook for partner review.

Output: docs/native-app-iap-analysis.xlsx
Tab 1: Parameters & Summary (editable assumptions + side-by-side table)
Tabs 2-6: One per strategy (Baseline, Options A-D)

All derived values are live formulas — change any assumption on the
Parameters tab and the other tabs update.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "/Users/chrispatton/Coding for Dummies/Comic Tracker/docs/native-app-iap-analysis.xlsx"

# --- Styles ---
TITLE = Font(bold=True, size=16, color="FFFFFF")
H1 = Font(bold=True, size=12, color="FFFFFF")
H2 = Font(bold=True, size=11)
BOLD = Font(bold=True)
INPUT_FONT = Font(bold=True, color="0B5394")

TITLE_FILL = PatternFill("solid", fgColor="1E3A5F")
H1_FILL = PatternFill("solid", fgColor="4472C4")
H2_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
NET_FILL = PatternFill("solid", fgColor="C6EFCE")
NEG_FILL = PatternFill("solid", fgColor="FFD6D6")

THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = '"$"#,##0'
MONEY_NEG = '"$"#,##0;[Red]-"$"#,##0'
PCT = "0.0%"
INT = "#,##0"

def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_row(ws, row, values, font=None, fill=None, fmt=None, border=False, align=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        if font: c.font = font
        if fill: c.fill = fill
        if fmt: c.number_format = fmt
        if border: c.border = BORDER
        if align: c.alignment = align

wb = Workbook()

# =============================================================================
# TAB 1: Parameters & Summary
# =============================================================================
ws = wb.active
ws.title = "Parameters & Summary"
set_widths(ws, [36, 16, 16, 16, 16, 16, 16, 16])

# Title
ws.cell(row=1, column=1, value="Native App IAP Analysis — Partner Review").font = TITLE
ws.cell(row=1, column=1).fill = TITLE_FILL
ws.merge_cells("A1:H1")
ws.row_dimensions[1].height = 28
ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

ws.cell(row=2, column=1, value="Edit yellow cells to model different scenarios. All other tabs update automatically.").font = Font(italic=True, color="666666")
ws.merge_cells("A2:H2")

# ---- Section 1: Assumptions ----
ws.cell(row=4, column=1, value="ASSUMPTIONS (edit yellow cells)").font = H1
ws.cell(row=4, column=1).fill = H1_FILL
ws.merge_cells("A4:H4")

ws.cell(row=5, column=1, value="Parameter").font = H2
ws.cell(row=5, column=2, value="Value").font = H2
ws.cell(row=5, column=3, value="Notes").font = H2
for col in range(1, 4):
    ws.cell(row=5, column=col).fill = H2_FILL

# User / sub counts
rows = [
    (6,  "Total users",                     10000, INT,   "Overall user base"),
    (7,  "Monthly subscribers",             3000,  INT,   "Paying $4.99/mo"),
    (8,  "Annual subscribers",              2000,  INT,   "Paying $49.99/yr"),
    (9,  "Scan pack buyers (annual)",       3000,  INT,   "$1.99/10 scans"),
    # Pricing
    (11, "Monthly sub price",               4.99,  MONEY, ""),
    (12, "Annual sub price",                49.99, MONEY, ""),
    (13, "Scan pack price",                 1.99,  MONEY, ""),
    # Platform split
    (15, "iOS share of premium users",      0.50,  PCT,   "% using iOS app"),
    (16, "Android share",                   0.40,  PCT,   "% using Android app"),
    (17, "Web share",                       0.10,  PCT,   "% web-only (always Stripe)"),
    # Fees
    (19, "Apple fee rate (SBP)",            0.15,  PCT,   "15% under $1M/yr dev revenue; 30% standard"),
    (20, "Google Play fee (SBP)",           0.15,  PCT,   "15% under $1M/yr"),
    (21, "Stripe percent fee",              0.029, PCT,   "2.9% per transaction"),
    (22, "Stripe flat fee per txn",         0.30,  MONEY, ""),
    # Conversion assumptions for C/D
    (24, "Opt C: iOS→web conversion",       0.60,  PCT,   "% of iOS users who'd pay who follow through when forced to web"),
    (25, "Opt D: iOS→web conversion",       0.20,  PCT,   "% of iOS users who give up premium vs switching to web"),
]

for row_num, label, value, fmt, note in rows:
    ws.cell(row=row_num, column=1, value=label)
    c = ws.cell(row=row_num, column=2, value=value)
    c.font = INPUT_FONT
    c.fill = INPUT_FILL
    c.number_format = fmt
    c.border = BORDER
    ws.cell(row=row_num, column=3, value=note).font = Font(italic=True, color="666666")

# Parameter cell references for use in formulas elsewhere
P = {
    "monthly_users": "'Parameters & Summary'!$B$7",
    "annual_users":  "'Parameters & Summary'!$B$8",
    "scan_buyers":   "'Parameters & Summary'!$B$9",
    "price_monthly": "'Parameters & Summary'!$B$11",
    "price_annual":  "'Parameters & Summary'!$B$12",
    "price_scan":    "'Parameters & Summary'!$B$13",
    "ios_share":     "'Parameters & Summary'!$B$15",
    "android_share": "'Parameters & Summary'!$B$16",
    "web_share":     "'Parameters & Summary'!$B$17",
    "apple_rate":    "'Parameters & Summary'!$B$19",
    "google_rate":   "'Parameters & Summary'!$B$20",
    "stripe_pct":    "'Parameters & Summary'!$B$21",
    "stripe_flat":   "'Parameters & Summary'!$B$22",
    "optc_conv":     "'Parameters & Summary'!$B$24",
    "optd_conv":     "'Parameters & Summary'!$B$25",
}

def gross_by_channel(share_ref):
    """Total gross revenue for a channel given its share reference."""
    return (f"({share_ref}*{P['monthly_users']}*{P['price_monthly']}*12)"
            f"+({share_ref}*{P['annual_users']}*{P['price_annual']})"
            f"+({share_ref}*{P['scan_buyers']}*{P['price_scan']})")

def stripe_fees_by_channel(share_ref):
    """Total Stripe fees for a channel if that channel is on Stripe."""
    return (
        f"({share_ref}*{P['monthly_users']}*12*({P['price_monthly']}*{P['stripe_pct']}+{P['stripe_flat']}))"
        f"+({share_ref}*{P['annual_users']}*({P['price_annual']}*{P['stripe_pct']}+{P['stripe_flat']}))"
        f"+({share_ref}*{P['scan_buyers']}*({P['price_scan']}*{P['stripe_pct']}+{P['stripe_flat']}))"
    )

# ---- Section 2: Channel Gross ----
ws.cell(row=28, column=1, value="CHANNEL GROSS REVENUE (derived)").font = H1
ws.cell(row=28, column=1).fill = H1_FILL
ws.merge_cells("A28:H28")

headers = ["Channel", "Share", "Monthly subs gross", "Annual subs gross", "Scan packs gross", "Total gross"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=29, column=i, value=h)
    c.font = H2
    c.fill = H2_FILL

channels = [
    ("iOS",     P["ios_share"],     30),
    ("Android", P["android_share"], 31),
    ("Web",     P["web_share"],     32),
]

for name, share_ref, row in channels:
    ws.cell(row=row, column=1, value=name).font = BOLD
    ws.cell(row=row, column=2, value=f"={share_ref}").number_format = PCT
    ws.cell(row=row, column=3, value=f"={share_ref}*{P['monthly_users']}*{P['price_monthly']}*12").number_format = MONEY
    ws.cell(row=row, column=4, value=f"={share_ref}*{P['annual_users']}*{P['price_annual']}").number_format = MONEY
    ws.cell(row=row, column=5, value=f"={share_ref}*{P['scan_buyers']}*{P['price_scan']}").number_format = MONEY
    ws.cell(row=row, column=6, value=f"=SUM(C{row}:E{row})").number_format = MONEY

# Total row
ws.cell(row=33, column=1, value="Total").font = BOLD
ws.cell(row=33, column=2, value="=SUM(B30:B32)").number_format = PCT
for col in range(3, 7):
    letter = get_column_letter(col)
    ws.cell(row=33, column=col, value=f"=SUM({letter}30:{letter}32)").number_format = MONEY
for col in range(1, 7):
    ws.cell(row=33, column=col).fill = TOTAL_FILL
    ws.cell(row=33, column=col).font = BOLD

# Store references for re-use in strategies
IOS_GROSS     = "'Parameters & Summary'!$F$30"
ANDROID_GROSS = "'Parameters & Summary'!$F$31"
WEB_GROSS     = "'Parameters & Summary'!$F$32"
TOTAL_GROSS   = "'Parameters & Summary'!$F$33"

# Stripe fees reference for each channel (if on Stripe)
ws.cell(row=35, column=1, value="STRIPE FEES PER CHANNEL (if channel is on Stripe)").font = H1
ws.cell(row=35, column=1).fill = H1_FILL
ws.merge_cells("A35:H35")
ws.cell(row=36, column=1, value="Channel").font = H2
ws.cell(row=36, column=2, value="Stripe fees").font = H2
ws.cell(row=36, column=1).fill = H2_FILL
ws.cell(row=36, column=2).fill = H2_FILL

ws.cell(row=37, column=1, value="iOS").font = BOLD
ws.cell(row=37, column=2, value=f"={stripe_fees_by_channel(P['ios_share'])}").number_format = MONEY
ws.cell(row=38, column=1, value="Android").font = BOLD
ws.cell(row=38, column=2, value=f"={stripe_fees_by_channel(P['android_share'])}").number_format = MONEY
ws.cell(row=39, column=1, value="Web").font = BOLD
ws.cell(row=39, column=2, value=f"={stripe_fees_by_channel(P['web_share'])}").number_format = MONEY

IOS_STRIPE     = "'Parameters & Summary'!$B$37"
ANDROID_STRIPE = "'Parameters & Summary'!$B$38"
WEB_STRIPE     = "'Parameters & Summary'!$B$39"

# ---- Section 3: Side-by-side ----
ws.cell(row=42, column=1, value="SIDE-BY-SIDE STRATEGY COMPARISON").font = H1
ws.cell(row=42, column=1).fill = H1_FILL
ws.merge_cells("A42:H42")

sbs_headers = ["Strategy", "Apple fee", "Google fee", "Stripe fees", "Total fees", "Net revenue", "vs Baseline", "% hit to net"]
for i, h in enumerate(sbs_headers, start=1):
    c = ws.cell(row=43, column=i, value=h)
    c.font = H2
    c.fill = H2_FILL

# Baseline (all Stripe)
ws.cell(row=44, column=1, value="Baseline (all Stripe, web only)").font = BOLD
ws.cell(row=44, column=2, value=0).number_format = MONEY
ws.cell(row=44, column=3, value=0).number_format = MONEY
ws.cell(row=44, column=4, value=f"={IOS_STRIPE}+{ANDROID_STRIPE}+{WEB_STRIPE}").number_format = MONEY
ws.cell(row=44, column=5, value="=B44+C44+D44").number_format = MONEY
ws.cell(row=44, column=6, value=f"={TOTAL_GROSS}-E44").number_format = MONEY
ws.cell(row=44, column=7, value="=F44-$F$44").number_format = MONEY_NEG
ws.cell(row=44, column=8, value="=(F44-$F$44)/$F$44").number_format = PCT

# Option A: Apple IAP + Stripe (Android+Web)
ws.cell(row=45, column=1, value="A — Apple IAP + Stripe (Android+Web)").font = BOLD
ws.cell(row=45, column=2, value=f"={IOS_GROSS}*{P['apple_rate']}").number_format = MONEY
ws.cell(row=45, column=3, value=0).number_format = MONEY
ws.cell(row=45, column=4, value=f"={ANDROID_STRIPE}+{WEB_STRIPE}").number_format = MONEY
ws.cell(row=45, column=5, value="=B45+C45+D45").number_format = MONEY
ws.cell(row=45, column=6, value=f"={TOTAL_GROSS}-E45").number_format = MONEY
ws.cell(row=45, column=7, value="=F45-$F$44").number_format = MONEY_NEG
ws.cell(row=45, column=8, value="=(F45-$F$44)/$F$44").number_format = PCT

# Option B: Apple IAP + Google Play + Stripe (Web)
ws.cell(row=46, column=1, value="B — Apple IAP + Google Play (Web on Stripe)").font = BOLD
ws.cell(row=46, column=2, value=f"={IOS_GROSS}*{P['apple_rate']}").number_format = MONEY
ws.cell(row=46, column=3, value=f"={ANDROID_GROSS}*{P['google_rate']}").number_format = MONEY
ws.cell(row=46, column=4, value=f"={WEB_STRIPE}").number_format = MONEY
ws.cell(row=46, column=5, value="=B46+C46+D46").number_format = MONEY
ws.cell(row=46, column=6, value=f"={TOTAL_GROSS}-E46").number_format = MONEY
ws.cell(row=46, column=7, value="=F46-$F$44").number_format = MONEY_NEG
ws.cell(row=46, column=8, value="=(F46-$F$44)/$F$44").number_format = PCT

# Option C: Web-only iOS, Stripe on Android
# Only conv% of iOS revenue is captured (on Stripe), rest lost
ws.cell(row=47, column=1, value="C — Web-only iOS upgrades").font = BOLD
ws.cell(row=47, column=2, value=0).number_format = MONEY
ws.cell(row=47, column=3, value=0).number_format = MONEY
ws.cell(row=47, column=4, value=f"=({IOS_STRIPE}*{P['optc_conv']})+{ANDROID_STRIPE}+{WEB_STRIPE}").number_format = MONEY
ws.cell(row=47, column=5, value="=B47+C47+D47").number_format = MONEY
# Net = (iOS captured × conv - iOS stripe × conv) + Android net + Web net
ws.cell(row=47, column=6, value=f"=({IOS_GROSS}*{P['optc_conv']})+{ANDROID_GROSS}+{WEB_GROSS}-E47").number_format = MONEY
ws.cell(row=47, column=7, value="=F47-$F$44").number_format = MONEY_NEG
ws.cell(row=47, column=8, value="=(F47-$F$44)/$F$44").number_format = PCT

# Option D: iOS free-only
ws.cell(row=48, column=1, value="D — iOS free-only (no iOS upgrades)").font = BOLD
ws.cell(row=48, column=2, value=0).number_format = MONEY
ws.cell(row=48, column=3, value=0).number_format = MONEY
ws.cell(row=48, column=4, value=f"=({IOS_STRIPE}*{P['optd_conv']})+{ANDROID_STRIPE}+{WEB_STRIPE}").number_format = MONEY
ws.cell(row=48, column=5, value="=B48+C48+D48").number_format = MONEY
ws.cell(row=48, column=6, value=f"=({IOS_GROSS}*{P['optd_conv']})+{ANDROID_GROSS}+{WEB_GROSS}-E48").number_format = MONEY
ws.cell(row=48, column=7, value="=F48-$F$44").number_format = MONEY_NEG
ws.cell(row=48, column=8, value="=(F48-$F$44)/$F$44").number_format = PCT

# Highlight Net column
for r in range(44, 49):
    ws.cell(row=r, column=6).fill = NET_FILL
    ws.cell(row=r, column=6).font = BOLD

# ---- Key takeaways ----
ws.cell(row=51, column=1, value="KEY TAKEAWAYS").font = H1
ws.cell(row=51, column=1).fill = H1_FILL
ws.merge_cells("A51:H51")

takeaways = [
    "1. Apple's auction marketplace is unaffected — physical goods are exempt from IAP. Only Premium + scan packs trigger Apple/Google fees.",
    "2. Option A (dual rails) is the sweet spot — Apple takes ~4% of net, but the alternative (web-only) loses 20%+ to conversion drop.",
    "3. Small Business Program critical: 15% applies under $1M/yr dev revenue. Above that, Apple = 30% y1 subs + 30% consumables (15% on y2+ renewals).",
    "4. Scan packs are the worst IAP economics (always consumable = 30% standard) but small absolute dollars ($5,970 gross).",
    "5. Engineering cost of IAP: ~3–4 weeks upfront to build StoreKit/Play Billing + receipt validation + entitlement sync. Ongoing maintenance.",
    "6. Reader-app exception: unlikely to qualify because scanning is a core digital feature. Spotify/Netflix-style carveout doesn't fit Collectors Chest.",
    "7. EU (Digital Markets Act) now allows external billing on iOS for EU users — could reduce Apple cut for ~20% of international users at scale.",
]
for i, t in enumerate(takeaways):
    c = ws.cell(row=52 + i, column=1, value=t)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=52 + i, start_column=1, end_row=52 + i, end_column=8)
    ws.row_dimensions[52 + i].height = 28

ws.freeze_panes = "A4"

# =============================================================================
# HELPER: Build a per-strategy detail tab
# =============================================================================
def build_strategy_tab(title, strategy_desc, rows_data, assumptions_extra=None):
    """rows_data: list of tuples (channel, gross_formula, fee_label, fee_formula, net_formula)."""
    s = wb.create_sheet(title=title)
    set_widths(s, [30, 18, 18, 18, 18, 22])

    # Title
    s.cell(row=1, column=1, value=title).font = TITLE
    s.cell(row=1, column=1).fill = TITLE_FILL
    s.merge_cells("A1:F1")
    s.row_dimensions[1].height = 26
    s.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # Description
    s.cell(row=2, column=1, value=strategy_desc).font = Font(italic=True)
    s.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    s.merge_cells("A2:F2")
    s.row_dimensions[2].height = 42

    # Assumptions linked from Params tab
    s.cell(row=4, column=1, value="KEY PARAMETERS (linked from Parameters tab)").font = H1
    s.cell(row=4, column=1).fill = H1_FILL
    s.merge_cells("A4:F4")

    param_rows = [
        ("Monthly subs", P["monthly_users"], INT),
        ("Annual subs", P["annual_users"], INT),
        ("Scan pack buyers/yr", P["scan_buyers"], INT),
        ("iOS share", P["ios_share"], PCT),
        ("Android share", P["android_share"], PCT),
        ("Web share", P["web_share"], PCT),
        ("Apple fee (SBP)", P["apple_rate"], PCT),
        ("Google fee (SBP)", P["google_rate"], PCT),
    ]
    if assumptions_extra:
        param_rows.extend(assumptions_extra)

    for i, (label, ref, fmt) in enumerate(param_rows):
        r = 5 + i
        s.cell(row=r, column=1, value=label).font = BOLD
        c = s.cell(row=r, column=2, value=f"={ref}")
        c.number_format = fmt
        c.fill = PatternFill("solid", fgColor="F3F3F3")

    # Breakdown table
    start = 5 + len(param_rows) + 2
    s.cell(row=start, column=1, value="PER-CHANNEL BREAKDOWN").font = H1
    s.cell(row=start, column=1).fill = H1_FILL
    s.merge_cells(f"A{start}:F{start}")

    headers = ["Channel", "Gross", "Apple fee", "Google fee", "Stripe fee", "Net"]
    for i, h in enumerate(headers, start=1):
        c = s.cell(row=start + 1, column=i, value=h)
        c.font = H2
        c.fill = H2_FILL

    net_cells = []
    gross_cells = []
    apple_cells = []
    google_cells = []
    stripe_cells = []
    for i, row in enumerate(rows_data):
        r = start + 2 + i
        channel, gross_f, apple_f, google_f, stripe_f = row
        s.cell(row=r, column=1, value=channel).font = BOLD
        s.cell(row=r, column=2, value=gross_f).number_format = MONEY
        s.cell(row=r, column=3, value=apple_f).number_format = MONEY
        s.cell(row=r, column=4, value=google_f).number_format = MONEY
        s.cell(row=r, column=5, value=stripe_f).number_format = MONEY
        s.cell(row=r, column=6, value=f"=B{r}-C{r}-D{r}-E{r}").number_format = MONEY
        s.cell(row=r, column=6).fill = NET_FILL
        s.cell(row=r, column=6).font = BOLD
        gross_cells.append(f"B{r}")
        apple_cells.append(f"C{r}")
        google_cells.append(f"D{r}")
        stripe_cells.append(f"E{r}")
        net_cells.append(f"F{r}")

    # Totals
    total_row = start + 2 + len(rows_data)
    s.cell(row=total_row, column=1, value="Total").font = BOLD
    s.cell(row=total_row, column=2, value=f"=SUM({gross_cells[0]}:{gross_cells[-1]})").number_format = MONEY
    s.cell(row=total_row, column=3, value=f"=SUM({apple_cells[0]}:{apple_cells[-1]})").number_format = MONEY
    s.cell(row=total_row, column=4, value=f"=SUM({google_cells[0]}:{google_cells[-1]})").number_format = MONEY
    s.cell(row=total_row, column=5, value=f"=SUM({stripe_cells[0]}:{stripe_cells[-1]})").number_format = MONEY
    s.cell(row=total_row, column=6, value=f"=SUM({net_cells[0]}:{net_cells[-1]})").number_format = MONEY
    for col in range(1, 7):
        s.cell(row=total_row, column=col).fill = TOTAL_FILL
        s.cell(row=total_row, column=col).font = BOLD

    # vs Baseline row
    baseline_net = "'Parameters & Summary'!$F$44"
    delta_row = total_row + 2
    s.cell(row=delta_row, column=1, value="Net vs Baseline").font = BOLD
    s.cell(row=delta_row, column=2, value=f"=F{total_row}-{baseline_net}").number_format = MONEY_NEG
    s.cell(row=delta_row, column=3, value="% of baseline net").font = Font(italic=True)
    s.cell(row=delta_row, column=4, value=f"=(F{total_row}-{baseline_net})/{baseline_net}").number_format = PCT

    s.freeze_panes = "A4"
    return s

# =============================================================================
# TAB 2: Baseline
# =============================================================================
build_strategy_tab(
    "Baseline (all Stripe)",
    "Current state — web-only, all payments via Stripe. No native apps. No Apple/Google fees. Reference point for all other scenarios.",
    [
        ("iOS (web)", f"={IOS_GROSS}",     0, 0, f"={IOS_STRIPE}"),
        ("Android (web)", f"={ANDROID_GROSS}", 0, 0, f"={ANDROID_STRIPE}"),
        ("Web",     f"={WEB_GROSS}",     0, 0, f"={WEB_STRIPE}"),
    ],
)

# =============================================================================
# TAB 3: Option A
# =============================================================================
build_strategy_tab(
    "Option A - Apple IAP + Stripe",
    "Ship iOS with Apple IAP for Premium + scan packs. Android + Web continue using Stripe (assumes Google Play User Choice Billing). Preserves full iOS revenue minus Apple's 15% SBP cut.",
    [
        ("iOS (Apple IAP)", f"={IOS_GROSS}", f"={IOS_GROSS}*{P['apple_rate']}", 0, 0),
        ("Android (Stripe)", f"={ANDROID_GROSS}", 0, 0, f"={ANDROID_STRIPE}"),
        ("Web (Stripe)", f"={WEB_GROSS}", 0, 0, f"={WEB_STRIPE}"),
    ],
)

# =============================================================================
# TAB 4: Option B
# =============================================================================
build_strategy_tab(
    "Option B - Both store billing",
    "Ship iOS with Apple IAP AND Android with Google Play Billing. Web stays on Stripe. Safest / most compliant, but both stores take a cut of digital subs and consumables.",
    [
        ("iOS (Apple IAP)", f"={IOS_GROSS}", f"={IOS_GROSS}*{P['apple_rate']}", 0, 0),
        ("Android (Google Play)", f"={ANDROID_GROSS}", 0, f"={ANDROID_GROSS}*{P['google_rate']}", 0),
        ("Web (Stripe)", f"={WEB_GROSS}", 0, 0, f"={WEB_STRIPE}"),
    ],
)

# =============================================================================
# TAB 5: Option C
# =============================================================================
build_strategy_tab(
    "Option C - Web-only upgrades",
    "iOS app has NO in-app purchase path. Users must sign up / upgrade on the website. Apple's anti-steering rules prohibit linking out. Avoids IAP fees but loses revenue to conversion drop (industry typical ~40% drop).",
    [
        ("iOS (web, partial capture)",
         f"={IOS_GROSS}*{P['optc_conv']}", 0, 0,
         f"={IOS_STRIPE}*{P['optc_conv']}"),
        ("Android (Stripe)", f"={ANDROID_GROSS}", 0, 0, f"={ANDROID_STRIPE}"),
        ("Web (Stripe)", f"={WEB_GROSS}", 0, 0, f"={WEB_STRIPE}"),
    ],
    assumptions_extra=[("Opt C conversion rate", P["optc_conv"], PCT)],
)

# =============================================================================
# TAB 6: Option D
# =============================================================================
build_strategy_tab(
    "Option D - iOS free-only",
    "iOS app has no upgrade path at all. Only free-tier features available on iOS. Users who want Premium must use web/Android. Lowest App Store rejection risk but caps iOS revenue severely.",
    [
        ("iOS (few go to web)",
         f"={IOS_GROSS}*{P['optd_conv']}", 0, 0,
         f"={IOS_STRIPE}*{P['optd_conv']}"),
        ("Android (Stripe)", f"={ANDROID_GROSS}", 0, 0, f"={ANDROID_STRIPE}"),
        ("Web (Stripe)", f"={WEB_GROSS}", 0, 0, f"={WEB_STRIPE}"),
    ],
    assumptions_extra=[("Opt D conversion rate", P["optd_conv"], PCT)],
)

# =============================================================================
# TAB 7: Growth Sensitivity
# =============================================================================
gs = wb.create_sheet(title="Growth Sensitivity")
set_widths(gs, [16, 16, 18, 20, 20, 20, 18, 18])

# References to baseline nets from side-by-side on Parameters tab
PWA_NET = "'Parameters & Summary'!$F$44"   # Baseline row
OPTA_NET = "'Parameters & Summary'!$F$45"  # Option A row
OPTB_NET = "'Parameters & Summary'!$F$46"  # Option B row

# Title
gs.cell(row=1, column=1, value="Growth Sensitivity — When Does Native Pay Off?").font = TITLE
gs.cell(row=1, column=1).fill = TITLE_FILL
gs.merge_cells("A1:H1")
gs.row_dimensions[1].height = 26
gs.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

gs.cell(row=2, column=1, value=(
    "Assumption: PWA-only revenue is capped at current user count (App Store is a growth channel the PWA doesn't tap). "
    "Native apps scale linearly with users if premium conversion % stays constant. "
    "Break-even: how much user growth native must bring to match PWA-only net."
)).font = Font(italic=True, color="666666")
gs.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
gs.merge_cells("A2:H2")
gs.row_dimensions[2].height = 56

# ---- Break-even analysis ----
gs.cell(row=4, column=1, value="BREAK-EVEN ANALYSIS").font = H1
gs.cell(row=4, column=1).fill = H1_FILL
gs.merge_cells("A4:H4")

gs.cell(row=5, column=1, value="Strategy").font = H2
gs.cell(row=5, column=2, value="Net at 1x users").font = H2
gs.cell(row=5, column=3, value="Break-even multiplier").font = H2
gs.cell(row=5, column=4, value="Extra users needed").font = H2
gs.cell(row=5, column=5, value="Interpretation").font = H2
for col in range(1, 6):
    gs.cell(row=5, column=col).fill = H2_FILL

gs.cell(row=6, column=1, value="Option A (Apple IAP + Stripe)").font = BOLD
gs.cell(row=6, column=2, value=f"={OPTA_NET}").number_format = MONEY
gs.cell(row=6, column=3, value=f"={PWA_NET}/{OPTA_NET}").number_format = "0.000"
gs.cell(row=6, column=4, value=f"=({PWA_NET}/{OPTA_NET}-1)*{P['monthly_users']}/0.3").number_format = INT
gs.cell(row=6, column=5, value="Extra users the App Store must bring just to offset Apple's cut")

gs.cell(row=7, column=1, value="Option B (Apple IAP + Google Play)").font = BOLD
gs.cell(row=7, column=2, value=f"={OPTB_NET}").number_format = MONEY
gs.cell(row=7, column=3, value=f"={PWA_NET}/{OPTB_NET}").number_format = "0.000"
gs.cell(row=7, column=4, value=f"=({PWA_NET}/{OPTB_NET}-1)*{P['monthly_users']}/0.3").number_format = INT
gs.cell(row=7, column=5, value="Extra users both stores' cuts must be offset by")

# ---- Sensitivity table ----
gs.cell(row=10, column=1, value="NET REVENUE AT DIFFERENT GROWTH MULTIPLIERS").font = H1
gs.cell(row=10, column=1).fill = H1_FILL
gs.merge_cells("A10:H10")

headers = [
    "User multiplier",
    "Total users",
    "Total premium",
    "PWA-only net",
    "Option A net",
    "Option B net",
    "A vs PWA",
    "B vs PWA",
]
for i, h in enumerate(headers, start=1):
    c = gs.cell(row=11, column=i, value=h)
    c.font = H2
    c.fill = H2_FILL

multipliers = [1.0, 1.1, 1.25, 1.5, 1.75, 2.0, 2.5, 3.0, 4.0, 5.0]

# Compute premium users as sum of all paying cohorts (for display)
premium_at_1x = f"({P['monthly_users']}+{P['annual_users']}+{P['scan_buyers']})"

for i, m in enumerate(multipliers):
    r = 12 + i
    gs.cell(row=r, column=1, value=m).number_format = "0.00\"x\""
    gs.cell(row=r, column=1).font = BOLD
    # Total users = base total × multiplier
    gs.cell(row=r, column=2, value=(
        f"=({P['monthly_users']}+{P['annual_users']})*(1/(('Parameters & Summary'!$B$7+'Parameters & Summary'!$B$8)/'Parameters & Summary'!$B$6))*A{r}"
    )).number_format = INT
    # Total paying cohorts scaled
    gs.cell(row=r, column=3, value=f"={premium_at_1x}*A{r}").number_format = INT
    # PWA-only net — FIXED at 1x (growth not available without App Store)
    gs.cell(row=r, column=4, value=f"={PWA_NET}").number_format = MONEY
    # Option A net scales linearly with users
    gs.cell(row=r, column=5, value=f"={OPTA_NET}*A{r}").number_format = MONEY
    # Option B net scales linearly with users
    gs.cell(row=r, column=6, value=f"={OPTB_NET}*A{r}").number_format = MONEY
    # Deltas
    gs.cell(row=r, column=7, value=f"=E{r}-D{r}").number_format = MONEY_NEG
    gs.cell(row=r, column=8, value=f"=F{r}-D{r}").number_format = MONEY_NEG

# Conditional highlight: color Net columns based on break-even
# Keep it simple — color the delta columns red/green depending on sign
from openpyxl.formatting.rule import CellIsRule
red_font = Font(color="9C0006")
green_font = Font(color="006100")
red_fill = PatternFill("solid", fgColor="FFC7CE")
green_fill = PatternFill("solid", fgColor="C6EFCE")

last_row = 11 + len(multipliers)
for col_letter in ("G", "H"):
    rng = f"{col_letter}12:{col_letter}{last_row}"
    gs.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=red_font))
    gs.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=green_fill, font=green_font))

# ---- Notes ----
notes_start = last_row + 3
gs.cell(row=notes_start, column=1, value="HOW TO READ THIS").font = H1
gs.cell(row=notes_start, column=1).fill = H1_FILL
gs.merge_cells(f"A{notes_start}:H{notes_start}")

notes = [
    "• PWA-only assumes the user base is capped at its current level. If word-of-mouth growth can double the PWA base too, this analysis understates PWA-only revenue.",
    "• Option A/B nets scale linearly with users, assuming premium conversion % (~50%) and platform split (50/40/10) stay constant.",
    "• Break-even shows the minimum user-base multiplier the App Store must deliver to match PWA-only net. Anything above break-even = native wins.",
    "• Typical rule of thumb: adding App Store + Play Store distribution grows consumer app user base 1.5x–3x over pure PWA/web (highly dependent on category and marketing).",
    "• Comics/collectibles apps tend toward the higher end of that range because app-store browsing is a common discovery path for niche-hobby apps.",
    "• If you expect <1.1x growth from native apps, Option A is essentially a wash — fees eat the incremental users. Above 1.5x, native clearly wins.",
    "• Alternative play: stay PWA-only for v1, add native apps once organic traction is proven and you can afford the IAP engineering upfront.",
]
for i, text in enumerate(notes):
    r = notes_start + 1 + i
    gs.cell(row=r, column=1, value=text).alignment = Alignment(wrap_text=True, vertical="top")
    gs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    gs.row_dimensions[r].height = 30

gs.freeze_panes = "A4"

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
